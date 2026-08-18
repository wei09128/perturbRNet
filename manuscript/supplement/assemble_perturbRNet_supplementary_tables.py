#!/usr/bin/env python3
"""Assemble copied perturbRNet result sources into a submission workbook.

Run after find_perturbRNet_supplementary_tables.py --copy.
The script never modifies source files. It creates:
  perturbRNet_supplementary_tables.xlsx
  supplementary_table_legends.md
  submission_panels/*.csv
"""

from __future__ import annotations

import argparse
import json
import re
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


TITLES = {
    "S1": "Simulation scenarios and performance",
    "S2": "Norman threshold and expression sensitivity",
    "S3": "Dataset, sample and analysis-design summary",
    "S4": "Alzheimer-model restoration candidates and calibration",
    "S5": "NASA comparison-level perturbRNet fits and null controls",
    "S6": "Norman held-out validation and robustness analyses",
}

LEGENDS = {
    "S1": "Simulation designs and performance summaries for end-to-end count generation, identifiability, estimator comparison and prior-sign uncertainty. Raw scenario-level outputs are distributed separately as machine-readable supplementary data.",
    "S2": "Sensitivity of the Norman hidden-cancellation endpoint and perturbRNet discrimination to response thresholds and gene expression. Extreme threshold cells with few positives are descriptive.",
    "S3": "Datasets, cohorts, samples and validation partitions used in the NASA, Alzheimer-model and Norman analyses. Norman cell partitions and gemgroups are technical partitions rather than biological replicates.",
    "S4": "Expression-restoration classifications, degree calibration and matched conditional-null results for AZD and shATM. Candidate genes are mechanistic hypotheses rather than validated causal mediators.",
    "S5": "NASA young-cortex and old-whole-brain comparison-level perturbRNet fits, identifier mapping and 200-draw destroyed-information controls. Cohorts were analyzed independently.",
    "S6": "Primary balance-constrained Norman validation and sensitivity analyses, including expression-aware matching, capture-partition cross-fitting and discovery-noise-standardized endpoints. Bootstrap resampling units were perturbation pairs.",
}


def token(value: str) -> str:
    return re.sub(r"[^a-z0-9]+", "_", value.lower()).strip("_")


def read_csv(path: Path) -> pd.DataFrame:
    return pd.read_csv(path, low_memory=False)


def json_frame(path: Path) -> pd.DataFrame:
    data = json.loads(path.read_text(encoding="utf-8"))
    rows = []
    for key, value in data.items():
        if isinstance(value, (list, dict)):
            value = json.dumps(value, ensure_ascii=False)
        rows.append({"field": key, "value": value})
    return pd.DataFrame(rows)


def find_one(root: Path, pattern: str, required: bool = True) -> Path | None:
    hits = sorted(root.glob(pattern))
    if not hits:
        if required:
            raise FileNotFoundError(f"Required source not found: {pattern}")
        return None
    if len(hits) > 1:
        raise RuntimeError(f"Ambiguous source pattern {pattern}: {hits}")
    return hits[0]


def panel(label: str, title: str, path: Path, frame: pd.DataFrame | None = None):
    return {
        "label": label,
        "title": title,
        "source": path.name,
        "data": read_csv(path) if frame is None else frame,
    }


def build_panels(root: Path):
    panels = {key: [] for key in TITLES}

    # S1: summary-level results only; scenario-level raw files remain archived.
    s1_specs = [
        ("A", "End-to-end method summary", "table_s1_*v3_1*method_summary.csv"),
        ("B", "End-to-end marginal scenario effects", "table_s1_*v3_1*marginal_summary.csv"),
        ("C", "Identifiability associations", "table_s1_*v4_1*associations.csv"),
        ("D", "Identifiability geometry by signal", "table_s1_*v4_1*geometry_by_signal.csv"),
        ("E", "Activity-estimator comparison", "table_s1_*method_comparison_summary.csv"),
        ("F", "Prior-sign uncertainty method summary", "table_s1_*uncertainty*method_summary.csv"),
        ("G", "Prior-sign uncertainty gain summary", "table_s1_*uncertainty*gain_summary.csv"),
    ]
    for label, title, pattern in s1_specs:
        path = find_one(root, pattern)
        panels["S1"].append(panel(label, title, path))

    # S2: all three finalized threshold/expression summaries.
    s2_specs = [
        ("A", "Threshold-grid discrimination", "table_s2_*threshold_expression_performance.csv"),
        ("B", "Threshold-grid matching coverage", "table_s2_*threshold_match_coverage.csv"),
        ("C", "Primary endpoint by expression quintile", "table_s2_*primary_endpoint_by_expression_quintile.csv"),
    ]
    for label, title, pattern in s2_specs:
        path = find_one(root, pattern)
        panels["S2"].append(panel(label, title, path))

    # S3: derive a compact NASA cohort table and retain Norman inventory/design metadata.
    samples = find_one(root, "table_s3_samples.csv")
    sample_df = read_csv(samples)
    sample_df["Group"] = sample_df["Group"].fillna("excluded_missing_group")
    cohort = (
        sample_df.groupby(["Tissue", "Group"], dropna=False)
        .size().rename("samples").reset_index()
        .sort_values(["Tissue", "Group"], kind="stable")
    )
    panels["S3"].append(panel("A", "NASA and Alzheimer-model sample inventory", samples, cohort))
    inv = find_one(root, "table_s3_*inventory_summary.json")
    panels["S3"].append(panel("B", "Norman dataset inventory", inv, json_frame(inv)))
    matrix = find_one(root, "table_s3_*matrix_diagnostics.csv")
    panels["S3"].append(panel("C", "Norman expression-matrix diagnostics", matrix))
    atlas = find_one(root, "table_s3_*atlas_manifest.json")
    panels["S3"].append(panel("D", "Norman response-atlas design", atlas, json_frame(atlas)))
    crossfit = find_one(root, "table_s3_*crossfit_manifest.json")
    panels["S3"].append(panel("E", "Norman discovery/validation design", crossfit, json_frame(crossfit)))

    # S4: classification, candidate lists, calibration and set-level null tests.
    s4_specs = [
        ("A", "Restoration-classification counts", "table_s4_*restoration_classification_counts.csv"),
        ("B", "AZD strong candidates", "table_s4_*azd_strong_candidates.csv"),
        ("C", "AZD low-amplitude exploratory candidates", "table_s4_*azd_low_amplitude_candidates.csv"),
        ("D", "shATM strong candidates", "table_s4_*shatm_strong_candidates.csv"),
        ("E", "shATM low-amplitude exploratory candidates", "table_s4_*shatm_low_amplitude_candidates.csv"),
        ("F", "Degree-association calibration", "table_s4_*degree_association_summary.csv"),
        ("G", "Matched conditional-null tests", "table_s4_*matched_set_test.csv"),
        ("H", "Restoration-threshold sensitivity", "table_s4_*threshold_sensitivity_summary.csv"),
        ("I", "Matched-set diagnostics", "table_s4_*matched_set_diagnostics.csv"),
    ]
    for label, title, pattern in s4_specs:
        path = find_one(root, pattern)
        panels["S4"].append(panel(label, title, path))

    # S5: finalized NASA outputs only.
    s5_specs = [
        ("A", "Comparison-level perturbRNet fits", "table_s5_nasa_prn_comparison_summary.csv"),
        ("B", "Identifier mapping and prior coverage", "table_s5_nasa_prn_mapping_summary.csv"),
        ("C", "Young-cortex 200-draw null controls", "table_s5_*young_nulls_final*null_comparison.csv"),
        ("D", "Old-whole-brain 200-draw null controls", "table_s5_*old_nulls_final*null_comparison.csv"),
    ]
    for label, title, pattern in s5_specs:
        path = find_one(root, pattern)
        panels["S5"].append(panel(label, title, path))

    # S6: final primary analysis first, then sensitivity analyses.
    s6_specs = [
        ("A", "Primary matching-specification balance grid", "table_s6_balance_constrained_matching_all_specification_balance.csv", True),
        ("B", "Primary selected matched discrimination", "table_s6_balance_constrained_matching_selected_benchmark_metrics.csv", True),
        ("C", "Primary perturbation-pair bootstrap", "table_s6_balance_constrained_matching_selected_pair_bootstrap.csv", True),
        ("D", "Primary minimum-positive sensitivity", "table_s6_balance_constrained_matching_minimum_positive_sensitivity.csv", False),
        ("E", "Expression-aware matched discrimination", "table_s6_expression_aware_matching_benchmark_metrics.csv", True),
        ("F", "Expression-aware covariate balance", "table_s6_expression_aware_matching_matching_balance.csv", True),
        ("G", "Fixed-threshold gemgroup forward discrimination", "table_s6_gemgroup_crossfit_forward*g1to4*selected_benchmark_metrics.csv", True),
        ("H", "Fixed-threshold gemgroup forward pair bootstrap", "table_s6_gemgroup_crossfit_forward*g1to4*selected_pair_bootstrap.csv", True),
        ("I", "Fixed-threshold gemgroup reverse discrimination", "table_s6_gemgroup_crossfit_reverse*g5to8*selected_benchmark_metrics.csv", True),
        ("J", "Fixed-threshold gemgroup reverse pair bootstrap", "table_s6_gemgroup_crossfit_reverse*g5to8*selected_pair_bootstrap.csv", True),
        ("K", "Noise-standardized forward discrimination", "table_s6_noise_standardized_crossfit_forward*g1to4*selected_benchmark_metrics.csv", True),
        ("L", "Noise-standardized forward pair bootstrap", "table_s6_noise_standardized_crossfit_forward*g1to4*selected_pair_bootstrap.csv", True),
        ("M", "Noise-standardized reverse discrimination", "table_s6_noise_standardized_crossfit_reverse*g5to8*selected_benchmark_metrics.csv", True),
        ("N", "Noise-standardized reverse pair bootstrap", "table_s6_noise_standardized_crossfit_reverse*g5to8*selected_pair_bootstrap.csv", True),
    ]
    for label, title, pattern, required in s6_specs:
        path = find_one(root, pattern, required=required)
        if path is not None:
            panels["S6"].append(panel(label, title, path))

    return panels


def write_panel_csvs(panels, out_dir: Path):
    out_dir.mkdir(parents=True, exist_ok=True)
    for table, items in panels.items():
        for item in items:
            name = f"table_{table.lower()}_{item['label'].lower()}_{token(item['title'])}.csv"
            item["data"].to_csv(out_dir / name, index=False)


def write_workbook(panels, output: Path):
    wb = Workbook()
    wb.remove(wb.active)
    dark = "1F4E78"
    light = "D9EAF7"

    readme = wb.create_sheet("Contents")
    readme.append(["perturbRNet supplementary tables"])
    readme.append(["Table", "Title", "Panels"])
    for table, title in TITLES.items():
        readme.append([table, title, len(panels[table])])
    readme["A1"].font = Font(bold=True, size=15, color="FFFFFF")
    readme["A1"].fill = PatternFill("solid", fgColor=dark)
    readme.merge_cells("A1:C1")
    readme.freeze_panes = "A3"
    readme.column_dimensions["A"].width = 12
    readme.column_dimensions["B"].width = 62
    readme.column_dimensions["C"].width = 12

    for table, items in panels.items():
        ws = wb.create_sheet(table)
        ws.sheet_view.showGridLines = False
        ws.cell(1, 1, f"Table {table}. {TITLES[table]}")
        ws.cell(1, 1).font = Font(bold=True, size=14, color="FFFFFF")
        ws.cell(1, 1).fill = PatternFill("solid", fgColor=dark)
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=12)
        ws.cell(2, 1, LEGENDS[table])
        ws.cell(2, 1).alignment = Alignment(wrap_text=True, vertical="top")
        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=12)
        ws.row_dimensions[2].height = 45
        row = 4
        for item in items:
            df = item["data"].copy()
            ws.cell(row, 1, f"Panel {item['label']}. {item['title']}")
            ws.cell(row, 1).font = Font(bold=True, color="FFFFFF")
            ws.cell(row, 1).fill = PatternFill("solid", fgColor=dark)
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=max(1, len(df.columns)))
            row += 1
            ws.cell(row, 1, f"Source: {item['source']}")
            ws.cell(row, 1).font = Font(italic=True, color="666666", size=9)
            row += 1
            for col_idx, name in enumerate(df.columns, 1):
                cell = ws.cell(row, col_idx, str(name))
                cell.font = Font(bold=True)
                cell.fill = PatternFill("solid", fgColor=light)
                cell.alignment = Alignment(wrap_text=True, vertical="top")
            row += 1
            for values in df.itertuples(index=False, name=None):
                for col_idx, value in enumerate(values, 1):
                    if pd.isna(value):
                        value = None
                    cell = ws.cell(row, col_idx, value)
                    cell.alignment = Alignment(vertical="top")
                    if isinstance(value, float):
                        cell.number_format = "0.0000"
                row += 1
            row += 2
        ws.freeze_panes = "A4"
        ws.auto_filter.ref = None
        for col in range(1, min(ws.max_column, 30) + 1):
            max_len = 0
            for r in range(1, min(ws.max_row, 400) + 1):
                value = ws.cell(r, col).value
                if value is not None:
                    max_len = max(max_len, len(str(value)))
            ws.column_dimensions[get_column_letter(col)].width = min(max(max_len + 2, 11), 34)

    output.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output)


def write_legends(path: Path):
    lines = ["# Supplementary table titles and legends", ""]
    for table in TITLES:
        lines.extend([
            f"## Table {table}. {TITLES[table]}", "", LEGENDS[table], ""
        ])
    path.write_text("\n".join(lines), encoding="utf-8")


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument(
        "--supplement-dir",
        type=Path,
        default=Path("/mnt/c/Wei/RRN/PRNet/manuscript/supplement"),
    )
    args = parser.parse_args()
    root = args.supplement_dir.resolve()
    if not root.is_dir():
        raise SystemExit(f"Supplement directory not found: {root}")

    panels = build_panels(root)
    workbook = root / "perturbRNet_supplementary_tables.xlsx"
    write_workbook(panels, workbook)
    write_panel_csvs(panels, root / "submission_panels")
    write_legends(root / "supplementary_table_legends.md")

    print(f"Wrote: {workbook}")
    print(f"Wrote: {root / 'supplementary_table_legends.md'}")
    print(f"Wrote panel CSVs: {root / 'submission_panels'}")
    for table in TITLES:
        print(f"  {table}: {len(panels[table])} panels")


if __name__ == "__main__":
    main()
